"""
DEMO COMPLETA DE ORDENAMIENTOS
Ejecuta y visualiza todos los métodos de ordenamiento (Internos y Externos)
Soporta carga de archivos: TXT y XLSX
"""

import tkinter as tk
from tkinter import ttk, messagebox
import random
import time
import threading
import re

# Importar las librerías creadas
from ordenamiento_interno import OrdenamientoInterno
from ordenamiento_externo import OrdenamientoExterno

# Intentar importar openpyxl para soporte Excel
try:
    from openpyxl import load_workbook
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    print("⚠️ openpyxl no instalado. Instálalo con: pip install openpyxl")


class DemoCompletaOrdenamiento:
    def __init__(self):
        self.ventana = tk.Tk()
        self.ventana.title("Demo Completa - Métodos de Ordenamiento")
        self.ventana.geometry("1300x750")
        self.ventana.configure(bg='#1a1a2e')
        
        # Datos
        self.datos_actuales = []
        self.animacion_activa = False
        self.nombre_archivo_cargado = ""
        
        # Métodos disponibles
        self.metodos_internos = {
            "Burbuja": OrdenamientoInterno.burbuja,
            "Inserción": OrdenamientoInterno.insercion,
            "Selección": OrdenamientoInterno.seleccion,
            "Shellsort": OrdenamientoInterno.shellsort,
            "Quicksort": OrdenamientoInterno.quicksort,
            "Heapsort": OrdenamientoInterno.heapsort,
            "Radixsort": OrdenamientoInterno.radixsort
        }
        
        self.metodos_externos = {
            "Intercalación": OrdenamientoExterno.intercalacion,
            "Mezcla Directa": OrdenamientoExterno.mezcla_directa,
            "Mezcla Equilibrada": OrdenamientoExterno.mezcla_equilibrada
        }
        
        self.setup_ui()
    
    def setup_ui(self):
        # Estilo
        style = ttk.Style()
        style.configure("TLabel", background="#1a1a2e", foreground="white")
        style.configure("TFrame", background="#1a1a2e")
        style.configure("TLabelframe", background="#1a1a2e", foreground="white")
        
        # Frame principal
        main_frame = ttk.Frame(self.ventana, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        titulo = tk.Label(main_frame, text="📊 MÉTODOS DE ORDENAMIENTO - COMPARATIVA VISUAL",
                         font=("Arial", 16, "bold"), bg="#1a1a2e", fg="#00d4ff")
        titulo.pack(pady=10)
        
        # Frame de datos
        datos_frame = ttk.LabelFrame(main_frame, text="Configuración de Datos", padding="10")
        datos_frame.pack(fill=tk.X, pady=5)
        
        # Botones de datos
        btn_frame = ttk.Frame(datos_frame)
        btn_frame.pack()
        
        tk.Button(btn_frame, text="🎲 Datos Aleatorios", command=self.generar_aleatorios,
                 bg="#27ae60", fg="white", font=("Arial", 10), padx=10, pady=5).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✏️ Ingresar Manual", command=self.ingresar_manual,
                 bg="#2980b9", fg="white", font=("Arial", 10), padx=10, pady=5).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="📁 Cargar TXT", command=self.cargar_archivo_txt,
                 bg="#8e44ad", fg="white", font=("Arial", 10), padx=10, pady=5).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="📊 Cargar Excel (XLSX)", command=self.cargar_archivo_excel,
                 bg="#e67e22", fg="white", font=("Arial", 10), padx=10, pady=5).pack(side=tk.LEFT, padx=5)
        
        # Mostrar datos actuales
        self.label_datos = tk.Label(datos_frame, text="Sin datos", bg="#1a1a2e", fg="#ecf0f1",
                                    font=("Arial", 10))
        self.label_datos.pack(pady=10)
        
        # Frame de métodos (scrollable)
        canvas_metodos = tk.Canvas(main_frame, bg="#1a1a2e", highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas_metodos.yview)
        self.frame_metodos = ttk.Frame(canvas_metodos)
        
        self.frame_metodos.bind("<Configure>", lambda e: canvas_metodos.configure(scrollregion=canvas_metodos.bbox("all")))
        canvas_metodos.create_window((0, 0), window=self.frame_metodos, anchor="nw")
        canvas_metodos.configure(yscrollcommand=scrollbar.set)
        
        canvas_metodos.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Crear frames para cada método
        self.frames_metodos = {}
        
        # Métodos Internos
        interno_frame = ttk.LabelFrame(self.frame_metodos, text="MÉTODOS INTERNOS", padding="10")
        interno_frame.pack(fill=tk.X, pady=5)
        
        for nombre in self.metodos_internos.keys():
            self._crear_frame_metodo(interno_frame, nombre)
        
        # Métodos Externos
        externo_frame = ttk.LabelFrame(self.frame_metodos, text="MÉTODOS EXTERNOS", padding="10")
        externo_frame.pack(fill=tk.X, pady=5)
        
        for nombre in self.metodos_externos.keys():
            self._crear_frame_metodo(externo_frame, nombre)
        
        # Frame de control general
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=10)
        
        tk.Button(control_frame, text="▶️ Ejecutar Todos", command=self.ejecutar_todos,
                 bg="#e74c3c", fg="white", font=("Arial", 11, "bold"), padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        tk.Button(control_frame, text="⏹️ Detener", command=self.detener_todos,
                 bg="#7f8c8d", fg="white", font=("Arial", 11), padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        tk.Button(control_frame, text="📈 Comparar Rendimiento", command=self.comparar_rendimiento,
                 bg="#3498db", fg="white", font=("Arial", 11), padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        tk.Button(control_frame, text="💾 Guardar Resultados", command=self.guardar_resultados,
                 bg="#2ecc71", fg="white", font=("Arial", 11), padx=20, pady=8).pack(side=tk.LEFT, padx=5)
        
        # Barra de estado
        self.status_label = tk.Label(main_frame, text="✅ Listo", bg="#1a1a2e", fg="#2ecc71",
                                     font=("Arial", 9))
        self.status_label.pack(pady=5)
    
    def _crear_frame_metodo(self, parent, nombre):
        """Crea un frame para un método de ordenamiento"""
        frame = tk.Frame(parent, bg="#16213e", relief=tk.RAISED, bd=1)
        frame.pack(fill=tk.X, pady=3)
        
        # Título del método
        label = tk.Label(frame, text=nombre, font=("Arial", 10, "bold"),
                        bg="#16213e", fg="#00d4ff", width=20, anchor="w")
        label.pack(side=tk.LEFT, padx=10, pady=5)
        
        # Canvas para visualización
        canvas = tk.Canvas(frame, bg="#0f3460", height=80, width=500)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Label para tiempo
        tiempo_label = tk.Label(frame, text="-- s", font=("Arial", 9),
                               bg="#16213e", fg="#f39c12", width=10)
        tiempo_label.pack(side=tk.RIGHT, padx=10)
        
        # Botón ejecutar individual
        btn = tk.Button(frame, text="Ejecutar", command=lambda n=nombre: self.ejecutar_individual(n),
                       bg="#2c3e50", fg="white", font=("Arial", 8), padx=10)
        btn.pack(side=tk.RIGHT, padx=5)
        
        self.frames_metodos[nombre] = {
            'frame': frame,
            'canvas': canvas,
            'tiempo_label': tiempo_label,
            'btn': btn,
            'datos_actuales': []
        }
    
    def generar_aleatorios(self):
        """Genera datos aleatorios"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Generar Datos")
        dialog.geometry("300x150")
        dialog.configure(bg="#1a1a2e")
        
        tk.Label(dialog, text="Número de elementos (5-30):", bg="#1a1a2e", fg="white").pack(pady=10)
        
        var = tk.IntVar(value=15)
        entry = tk.Entry(dialog, textvariable=var)
        entry.pack(pady=5)
        
        def aceptar():
            try:
                n = var.get()
                if 5 <= n <= 30:
                    self.datos_actuales = [random.randint(1, 100) for _ in range(n)]
                    self.label_datos.config(text=f"Datos: {self.datos_actuales}")
                    self.nombre_archivo_cargado = ""
                    self.status_label.config(text=f"✅ Generados {n} datos aleatorios")
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "Ingrese un número entre 5 y 30")
            except:
                messagebox.showerror("Error", "Ingrese un número válido")
        
        tk.Button(dialog, text="Aceptar", command=aceptar, bg="#27ae60", fg="white").pack(pady=10)
    
    def ingresar_manual(self):
        """Ingreso manual de datos"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Ingresar Datos")
        dialog.geometry("400x200")
        dialog.configure(bg="#1a1a2e")
        
        tk.Label(dialog, text="Ingrese números separados por espacios o comas:", 
                bg="#1a1a2e", fg="white").pack(pady=10)
        
        entry = tk.Entry(dialog, width=50)
        entry.pack(pady=10)
        
        def aceptar():
            texto = entry.get()
            texto = texto.replace(',', ' ')
            numeros = texto.split()
            try:
                datos = [int(n) for n in numeros]
                if 1 <= len(datos) <= 30:
                    self.datos_actuales = datos
                    self.label_datos.config(text=f"Datos: {self.datos_actuales}")
                    self.nombre_archivo_cargado = ""
                    self.status_label.config(text=f"✅ Ingresados {len(datos)} datos manualmente")
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "Ingrese entre 1 y 30 números")
            except:
                messagebox.showerror("Error", "Ingrese solo números válidos")
        
        tk.Button(dialog, text="Aceptar", command=aceptar, bg="#27ae60", fg="white").pack(pady=10)
    
    def cargar_archivo_txt(self):
        """Carga datos desde archivo de texto"""
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo TXT",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    contenido = f.read()
                
                # Extraer números
                numeros = re.findall(r'-?\d+', contenido)
                datos = [int(n) for n in numeros]
                
                if datos:
                    if len(datos) > 50:
                        if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos)} números. ¿Continuar?"):
                            return
                    
                    self.datos_actuales = datos[:50]
                    self.nombre_archivo_cargado = archivo
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)}): {self.datos_actuales[:20]}{'...' if len(datos) > 20 else ''}")
                    self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde {archivo.split('/')[-1]}")
                else:
                    messagebox.showerror("Error", "No se encontraron números en el archivo")
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo: {str(e)}")
    
    def cargar_archivo_excel(self):
        """Carga datos desde archivo Excel (.xlsx)"""
        if not EXCEL_DISPONIBLE:
            messagebox.showerror("Error", 
                               "openpyxl no está instalado.\n\n"
                               "Instálalo con el comando:\n"
                               "pip install openpyxl")
            return
        
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                # Cargar el libro de Excel
                wb = load_workbook(archivo, data_only=True)
                
                # Mostrar hojas disponibles para seleccionar
                hojas = wb.sheetnames
                
                if not hojas:
                    messagebox.showerror("Error", "El archivo no contiene hojas")
                    return
                
                # Diálogo para seleccionar hoja y columna
                self._seleccionar_hoja_excel(archivo, wb, hojas)
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo Excel:\n{str(e)}")
    
    def _seleccionar_hoja_excel(self, archivo, wb, hojas):
        """Diálogo para seleccionar hoja y columna del Excel"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Seleccionar datos del Excel")
        dialog.geometry("500x400")
        dialog.configure(bg="#1a1a2e")
        
        tk.Label(dialog, text="Seleccione la hoja:", bg="#1a1a2e", fg="white",
                font=("Arial", 10, "bold")).pack(pady=10)
        
        # Combobox para hojas
        hoja_var = tk.StringVar(value=hojas[0])
        hoja_combo = ttk.Combobox(dialog, textvariable=hoja_var, values=hojas, width=40)
        hoja_combo.pack(pady=5)
        
        # Frame para vista previa
        tk.Label(dialog, text="Vista previa de la hoja:", bg="#1a1a2e", fg="white",
                font=("Arial", 10, "bold")).pack(pady=10)
        
        preview_frame = tk.Frame(dialog, bg="#0f3460")
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        preview_text = tk.Text(preview_frame, bg="#0f3460", fg="white", height=10, width=60)
        preview_text.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar para preview
        scroll_preview = ttk.Scrollbar(preview_text, command=preview_text.yview)
        preview_text.configure(yscrollcommand=scroll_preview.set)
        
        def actualizar_preview(*args):
            """Actualiza la vista previa al cambiar de hoja"""
            preview_text.delete(1.0, tk.END)
            hoja_actual = hoja_var.get()
            ws = wb[hoja_actual]
            
            # Mostrar primeras filas
            preview_text.insert(tk.END, f"Hoja: {hoja_actual}\n")
            preview_text.insert(tk.END, "-" * 50 + "\n")
            
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i > 20:  # Limitar a 20 filas
                    preview_text.insert(tk.END, "...\n")
                    break
                # Filtrar solo valores numéricos
                valores = [str(cell) if cell is not None else "" for cell in row]
                preview_text.insert(tk.END, f"Fila {i}: {valores}\n")
        
        hoja_var.trace('w', actualizar_preview)
        actualizar_preview()
        
        # Opciones de columna
        tk.Label(dialog, text="Seleccione la columna:", bg="#1a1a2e", fg="white",
                font=("Arial", 10, "bold")).pack(pady=5)
        
        columna_var = tk.StringVar(value="A")
        columna_entry = tk.Entry(dialog, textvariable=columna_var, width=10, justify="center")
        columna_entry.pack(pady=5)
        
        tk.Label(dialog, text="(Ejemplo: A, B, C, etc.)", bg="#1a1a2e", fg="#7f8c8d",
                font=("Arial", 8)).pack()
        
        # Opciones adicionales
        incluir_encabezado = tk.BooleanVar(value=False)
        tk.Checkbutton(dialog, text="La primera fila es encabezado (ignorar)",
                      variable=incluir_encabezado, bg="#1a1a2e", fg="white",
                      selectcolor="#1a1a2e").pack(pady=5)
        
        def aceptar():
            try:
                hoja_seleccionada = hoja_var.get()
                columna = columna_var.get().upper()
                
                # Validar columna
                if not columna or len(columna) > 2:
                    messagebox.showerror("Error", "Columna inválida. Use letras (A, B, C, etc.)")
                    return
                
                # Convertir letra a índice (A=1, B=2, ...)
                col_idx = ord(columna[0]) - ord('A') + 1
                if len(columna) == 2:
                    col_idx = (ord(columna[0]) - ord('A') + 1) * 26 + (ord(columna[1]) - ord('A') + 1)
                
                ws = wb[hoja_seleccionada]
                datos = []
                
                start_row = 2 if incluir_encabezado.get() else 1
                
                for row in ws.iter_rows(min_row=start_row, values_only=True):
                    if len(row) >= col_idx:
                        valor = row[col_idx - 1]
                        if valor is not None:
                            try:
                                # Intentar convertir a número
                                num = float(valor) if isinstance(valor, (int, float)) else int(float(valor))
                                datos.append(int(num))
                            except (ValueError, TypeError):
                                pass  # Ignorar valores no numéricos
                
                if not datos:
                    messagebox.showerror("Error", "No se encontraron números en la columna seleccionada")
                    return
                
                if len(datos) > 50:
                    if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos)} números. ¿Continuar?"):
                        return
                
                self.datos_actuales = datos[:50]
                self.nombre_archivo_cargado = archivo
                self.label_datos.config(text=f"Datos Excel ({len(self.datos_actuales)}): {self.datos_actuales[:20]}{'...' if len(datos) > 20 else ''}")
                self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde Excel (Hoja: {hoja_seleccionada}, Col: {columna})")
                dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar Excel:\n{str(e)}")
        
        btn_frame = tk.Frame(dialog, bg="#1a1a2e")
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Aceptar", command=aceptar, bg="#27ae60", fg="white",
                 padx=20, pady=5).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="Cancelar", command=dialog.destroy, bg="#e74c3c", fg="white",
                 padx=20, pady=5).pack(side=tk.LEFT, padx=10)
    
    def dibujar_barras(self, canvas, datos):
        """Dibuja gráfico de barras en un canvas"""
        canvas.delete("all")
        
        if not datos:
            return
        
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        
        if ancho < 10:
            ancho = 500
        if alto < 10:
            alto = 80
        
        n = len(datos)
        if n == 0:
            return
        
        ancho_barra = max(5, (ancho - 20) / n - 2)
        max_valor = max(datos) if datos else 100
        
        for i, valor in enumerate(datos):
            x0 = 10 + i * (ancho_barra + 2)
            y0 = alto - (valor / max_valor) * (alto - 20)
            x1 = x0 + ancho_barra
            y1 = alto - 10
            
            # Color basado en el valor (gradiente)
            r = int((valor / max_valor) * 255)
            b = int((1 - valor / max_valor) * 255)
            color = f'#{r:02x}40{b:02x}'
            canvas.create_rectangle(x0, y0, x1, y1, fill=color, outline="white", width=1)
    
    def ejecutar_individual(self, nombre):
        """Ejecuta un método específico"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "Primero genere o cargue datos")
            return
        
        # Obtener el método
        metodo = None
        if nombre in self.metodos_internos:
            metodo = self.metodos_internos[nombre]
        elif nombre in self.metodos_externos:
            metodo = self.metodos_externos[nombre]
        
        if not metodo:
            return
        
        # Deshabilitar botón mientras ejecuta
        frame_info = self.frames_metodos[nombre]
        frame_info['btn'].config(state=tk.DISABLED, text="Ejecutando...")
        
        def ejecutar():
            datos_copy = self.datos_actuales.copy()
            
            start = time.time()
            
            def actualizar_paso(datos_paso):
                self.dibujar_barras(frame_info['canvas'], datos_paso)
                frame_info['canvas'].update()
                time.sleep(0.02)
            
            _, resultado = metodo(datos_copy, actualizar_paso)
            tiempo = time.time() - start
            
            # Mostrar resultado final
            self.dibujar_barras(frame_info['canvas'], resultado)
            frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
            frame_info['btn'].config(state=tk.NORMAL, text="Ejecutar")
        
        thread = threading.Thread(target=ejecutar)
        thread.daemon = True
        thread.start()
    
    def ejecutar_todos(self):
        """Ejecuta todos los métodos simultáneamente"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "Primero genere o cargue datos")
            return
        
        self.status_label.config(text="🔄 Ejecutando todos los métodos...")
        
        def ejecutar():
            hilos = []
            
            for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
                frame_info = self.frames_metodos[nombre]
                frame_info['btn'].config(state=tk.DISABLED, text="Ejecutando...")
                
                metodo = None
                if nombre in self.metodos_internos:
                    metodo = self.metodos_internos[nombre]
                else:
                    metodo = self.metodos_externos[nombre]
                
                thread = threading.Thread(target=self._ejecutar_metodo_hilo, args=(nombre, metodo, frame_info))
                thread.daemon = True
                thread.start()
                hilos.append(thread)
            
            for h in hilos:
                h.join()
            
            self.status_label.config(text="✅ Todos los métodos completados")
        
        threading.Thread(target=ejecutar, daemon=True).start()
    
    def _ejecutar_metodo_hilo(self, nombre, metodo, frame_info):
        """Ejecuta un método en un hilo"""
        datos_copy = self.datos_actuales.copy()
        
        start = time.time()
        
        def actualizar_paso(datos_paso):
            self.dibujar_barras(frame_info['canvas'], datos_paso)
            frame_info['canvas'].update()
            time.sleep(0.01)
        
        _, resultado = metodo(datos_copy, actualizar_paso)
        tiempo = time.time() - start
        
        self.dibujar_barras(frame_info['canvas'], resultado)
        frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
        frame_info['btn'].config(state=tk.NORMAL, text="Ejecutar")
    
    def detener_todos(self):
        """Detiene la ejecución (no puede detener hilos directamente)"""
        self.status_label.config(text="⏹️ Detenido (la ejecución continuará hasta terminar)")
        messagebox.showinfo("Info", "Las ejecuciones en curso continuarán hasta finalizar")
    
    def comparar_rendimiento(self):
        """Compara el rendimiento de todos los métodos"""
        if not self.datos_actuales or len(self.datos_actuales) < 2:
            messagebox.showwarning("Advertencia", "Se necesitan al menos 2 elementos")
            return
        
        resultados = f"""
╔══════════════════════════════════════════════════════════════╗
║              COMPARACIÓN DE RENDIMIENTO                      ║
╠══════════════════════════════════════════════════════════════╣
║ Elementos: {len(self.datos_actuales):<46} ║
╠══════════════════════════════════════════════════════════════╣
"""
        
        # Medir tiempos para cada método
        for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
            metodo = None
            if nombre in self.metodos_internos:
                metodo = self.metodos_internos[nombre]
            else:
                metodo = self.metodos_externos[nombre]
            
            datos_copy = self.datos_actuales.copy()
            start = time.time()
            _, _ = metodo(datos_copy)
            tiempo = time.time() - start
            
            resultados += f"║ {nombre:<20} {tiempo:.6f} segundos{' ' * (25 - len(f'{tiempo:.6f}'))}║\n"
        
        resultados += "╚══════════════════════════════════════════════════════════════╝"
        
        # Mostrar resultados en ventana
        ventana_resultados = tk.Toplevel(self.ventana)
        ventana_resultados.title("Resultados de Rendimiento")
        ventana_resultados.geometry("650x550")
        ventana_resultados.configure(bg="#1a1a2e")
        
        text_area = tk.Text(ventana_resultados, bg="#0a0a1a", fg="#00ff00", font=("Courier", 10))
        text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_area.insert(tk.END, resultados)
        text_area.config(state=tk.DISABLED)
        
        tk.Button(ventana_resultados, text="Cerrar", command=ventana_resultados.destroy,
                 bg="#e74c3c", fg="white", padx=20, pady=5).pack(pady=10)
        
        self.status_label.config(text="📊 Comparación de rendimiento completada")
    
    def guardar_resultados(self):
        """Guarda los resultados ordenados en un archivo"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "No hay datos para guardar")
            return
        
        from tkinter import filedialog
        archivo = filedialog.asksaveasfilename(
            title="Guardar resultados",
            defaultextension=".txt",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                with open(archivo, 'w', encoding='utf-8') as f:
                    f.write("=" * 70 + "\n")
                    f.write("RESULTADOS DE ORDENAMIENTO\n")
                    f.write("=" * 70 + "\n\n")
                    f.write(f"Datos originales ({len(self.datos_actuales)} elementos):\n")
                    f.write(f"{self.datos_actuales}\n\n")
                    f.write("-" * 70 + "\n\n")
                    
                    # Guardar resultados de cada método
                    for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
                        metodo = None
                        if nombre in self.metodos_internos:
                            metodo = self.metodos_internos[nombre]
                        else:
                            metodo = self.metodos_externos[nombre]
                        
                        datos_copy = self.datos_actuales.copy()
                        _, resultado = metodo(datos_copy)
                        
                        f.write(f"Método: {nombre}\n")
                        f.write(f"Resultado: {resultado}\n")
                        f.write("-" * 50 + "\n\n")
                
                self.status_label.config(text=f"✅ Resultados guardados en {archivo.split('/')[-1]}")
                messagebox.showinfo("Éxito", "Resultados guardados correctamente")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")
    
    def ejecutar(self):
        """Inicia la aplicación"""
        self.ventana.mainloop()


if __name__ == "__main__":
    app = DemoCompletaOrdenamiento()
    app.ejecutar()