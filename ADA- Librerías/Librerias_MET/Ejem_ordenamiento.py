"""
DEMO COMPLETA DE ORDENAMIENTOS
Ejecuta y visualiza todos los métodos de ordenamiento (Internos y Externos)
Soporta carga de archivos: TXT, XLSX y JSON
"""

import tkinter as tk
from tkinter import ttk, messagebox
import random
import time
import threading
import re
import json
from datetime import datetime

# Importar las librerías creadas
from ordenamiento_interno import OrdenamientoInterno
from ordenamiento_externo import OrdenamientoExterno

# Intentar importar openpyxl para soporte Excel
try:
    from openpyxl import load_workbook
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False
    print("⚠️ openpyxl no instalado. Instálalo con: pip install openpyxl")


class DemoCompletaOrdenamiento:
    def __init__(self):
        self.ventana = tk.Tk()
        self.ventana.title("Demo Completa - Métodos de Ordenamiento")
        self.ventana.geometry("1400x850")
        self.ventana.configure(bg='#1a1a2e')
        
        # Datos
        self.datos_actuales = []
        self.animacion_activa = False
        self.nombre_archivo_cargado = ""
        self.modo_visualizacion = "grafico"
        
        # Almacenar últimos resultados
        self.ultimos_resultados = {}
        
        # Métodos disponibles
        self.metodos_internos = {
            "Burbuja": OrdenamientoInterno.burbuja,
            "Inserción": OrdenamientoInterno.insercion,
            "Selección": OrdenamientoInterno.seleccion,
            "Shellsort": OrdenamientoInterno.shellsort,
            "Quicksort": OrdenamientoInterno.quicksort,
            "Heapsort": OrdenamientoInterno.heapsort,
            "Radixsort": OrdenamientoInterno.radixsort
        }
        
        self.metodos_externos = {
            "Intercalación": OrdenamientoExterno.intercalacion,
            "Mezcla Directa": OrdenamientoExterno.mezcla_directa,
            "Mezcla Equilibrada": OrdenamientoExterno.mezcla_equilibrada
        }
        
        self.setup_ui()
    
    def setup_ui(self):
        # Estilo
        style = ttk.Style()
        style.configure("TLabel", background="#1a1a2e", foreground="white")
        style.configure("TFrame", background="#1a1a2e")
        style.configure("TLabelframe", background="#1a1a2e", foreground="white")
        
        # Frame principal con paneles divididos
        paned_window = tk.PanedWindow(self.ventana, orient=tk.VERTICAL, bg="#1a1a2e", sashrelief=tk.RAISED, sashwidth=5)
        paned_window.pack(fill=tk.BOTH, expand=True)
        
        # Panel superior para controles
        top_panel = tk.Frame(paned_window, bg="#1a1a2e", height=200)
        paned_window.add(top_panel, height=200)
        
        # Panel inferior para resultados (scrollable)
        bottom_panel = tk.Frame(paned_window, bg="#1a1a2e")
        paned_window.add(bottom_panel)
        
        # ========== PANEL SUPERIOR ==========
        # Título
        titulo = tk.Label(top_panel, text="📊 MÉTODOS DE ORDENAMIENTO - COMPARATIVA VISUAL",
                         font=("Arial", 16, "bold"), bg="#1a1a2e", fg="#00d4ff")
        titulo.pack(pady=5)
        
        # Frame de datos (primera fila)
        datos_frame = ttk.LabelFrame(top_panel, text="Configuración de Datos", padding="10")
        datos_frame.pack(fill=tk.X, pady=5, padx=10)
        
        # Botones de datos - Fila 1
        btn_frame1 = tk.Frame(datos_frame, bg="#1a1a2e")
        btn_frame1.pack(pady=5)
        
        tk.Button(btn_frame1, text="🎲 Datos Aleatorios", command=self.generar_aleatorios,
                 bg="#27ae60", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="✏️ Ingresar Manual", command=self.ingresar_manual,
                 bg="#2980b9", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="📁 Cargar TXT", command=self.cargar_archivo_txt,
                 bg="#8e44ad", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="📊 Cargar Excel", command=self.cargar_archivo_excel,
                 bg="#e67e22", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame1, text="📄 Cargar JSON", command=self.cargar_archivo_json,
                 bg="#f39c12", fg="white", font=("Arial", 10), padx=10, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        
        # Botones de control - Fila 2
        btn_frame2 = tk.Frame(datos_frame, bg="#1a1a2e")
        btn_frame2.pack(pady=5)
        
        tk.Button(btn_frame2, text="▶️ Ejecutar Todos", command=self.ejecutar_todos,
                 bg="#e74c3c", fg="white", font=("Arial", 10, "bold"), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="⏹️ Detener", command=self.detener_todos,
                 bg="#7f8c8d", fg="white", font=("Arial", 10), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="📈 Comparar Rendimiento", command=self.comparar_rendimiento,
                 bg="#3498db", fg="white", font=("Arial", 10), padx=15, pady=5, width=18).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="💾 Guardar Todo", command=self.guardar_todos_resultados,
                 bg="#2ecc71", fg="white", font=("Arial", 10), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame2, text="🔄 Limpiar Todo", command=self.limpiar_todo,
                 bg="#95a5a6", fg="white", font=("Arial", 10), padx=15, pady=5, width=15).pack(side=tk.LEFT, padx=5)
        
        # Mostrar datos actuales
        self.label_datos = tk.Label(datos_frame, text="Sin datos", bg="#1a1a2e", fg="#ecf0f1",
                                    font=("Arial", 10), wraplength=1200)
        self.label_datos.pack(pady=5)
        
        # Barra de estado
        self.status_label = tk.Label(top_panel, text="✅ Listo", bg="#1a1a2e", fg="#2ecc71",
                                     font=("Arial", 9))
        self.status_label.pack(pady=5)
        
        # ========== PANEL INFERIOR ==========
        # Frame para scroll de métodos
        canvas_metodos = tk.Canvas(bottom_panel, bg="#1a1a2e", highlightthickness=0)
        scrollbar_vertical = ttk.Scrollbar(bottom_panel, orient="vertical", command=canvas_metodos.yview)
        scrollbar_horizontal = ttk.Scrollbar(bottom_panel, orient="horizontal", command=canvas_metodos.xview)
        
        self.frame_metodos = ttk.Frame(canvas_metodos)
        
        self.frame_metodos.bind("<Configure>", lambda e: canvas_metodos.configure(scrollregion=canvas_metodos.bbox("all")))
        canvas_metodos.create_window((0, 0), window=self.frame_metodos, anchor="nw")
        canvas_metodos.configure(yscrollcommand=scrollbar_vertical.set, xscrollcommand=scrollbar_horizontal.set)
        
        canvas_metodos.pack(side="left", fill="both", expand=True)
        scrollbar_vertical.pack(side="right", fill="y")
        scrollbar_horizontal.pack(side="bottom", fill="x")
        
        # Crear frames para cada método en grid
        self.frames_metodos = {}
        
        # Configurar grid con 4 columnas para mejor distribución
        for i in range(4):
            self.frame_metodos.grid_columnconfigure(i, weight=1)
        
        # Título de métodos internos
        interno_title = tk.Label(self.frame_metodos, text="🔹 MÉTODOS INTERNOS 🔹", 
                                 font=("Arial", 14, "bold"), bg="#1a1a2e", fg="#00d4ff")
        interno_title.grid(row=0, column=0, columnspan=4, pady=15)
        
        # Métodos Internos - organizados en 2 filas de 4 columnas
        # Fila 1: Burbuja, Inserción, Selección, Shellsort
        # Fila 2: Quicksort, Heapsort, Radixsort, (espacio)
        metodos_internos_lista = list(self.metodos_internos.keys())
        
        # Fila 1 - primeros 4 métodos
        row = 1
        for col, nombre in enumerate(metodos_internos_lista[:4]):
            self._crear_frame_metodo_grid(nombre, row, col)
        
        # Fila 2 - siguientes métodos (Quicksort, Heapsort, Radixsort)
        row = 2
        for col, nombre in enumerate(metodos_internos_lista[4:]):
            self._crear_frame_metodo_grid(nombre, row, col)
        
        # Título de métodos externos
        externo_title = tk.Label(self.frame_metodos, text="🔸 MÉTODOS EXTERNOS 🔸", 
                                 font=("Arial", 14, "bold"), bg="#1a1a2e", fg="#00d4ff")
        externo_title.grid(row=3, column=0, columnspan=4, pady=15)
        
        # Métodos Externos - centrados en 3 columnas
        row = 4
        metodos_externos_lista = list(self.metodos_externos.keys())
        
        # Calcular offset para centrar (3 métodos, centrarlos en 4 columnas)
        offset = (4 - len(metodos_externos_lista)) // 2
        
        for idx, nombre in enumerate(metodos_externos_lista):
            col = offset + idx
            self._crear_frame_metodo_grid(nombre, row, col)
        
        # Ajustar tamaño de los frames de métodos para mejor visualización
        self.ventana.update_idletasks()
        for nombre, info in self.frames_metodos.items():
            info['canvas'].config(width=300, height=140)
    
    def _crear_frame_metodo_grid(self, nombre, row, col):
        """Crea un frame para un método de ordenamiento en grid con tamaño mejorado"""
        frame = tk.Frame(self.frame_metodos, bg="#16213e", relief=tk.RAISED, bd=2, padx=8, pady=8)
        frame.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
        
        # Título del método con color especial para Quicksort y Radixsort
        if nombre == "Quicksort":
            titulo_color = "#e74c3c"
            metodo_destacado = True
        elif nombre == "Radixsort":
            titulo_color = "#f39c12"
            metodo_destacado = True
        else:
            titulo_color = "#00d4ff"
            metodo_destacado = False
        
        label = tk.Label(frame, text=nombre, font=("Arial", 11, "bold"),
                        bg="#16213e", fg=titulo_color)
        label.pack(pady=5)
        
        if metodo_destacado:
            # Marco de resaltado para métodos importantes
            highlight_frame = tk.Frame(frame, bg=titulo_color, height=2)
            highlight_frame.pack(fill=tk.X, padx=5)
        
        # Canvas para visualización (tamaño aumentado)
        canvas = tk.Canvas(frame, bg="#0f3460", height=140, width=300)
        canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Frame para botones
        botones_frame = tk.Frame(frame, bg="#16213e")
        botones_frame.pack(fill=tk.X, pady=5)
        
        # Label para tiempo
        tiempo_label = tk.Label(frame, text="-- s", font=("Arial", 9),
                               bg="#16213e", fg="#f39c12")
        tiempo_label.pack(pady=2)
        
        # Botón ejecutar individual
        btn_ejecutar = tk.Button(botones_frame, text="▶️ Ejecutar", 
                                command=lambda n=nombre: self.ejecutar_individual(n),
                                bg="#2c3e50", fg="white", font=("Arial", 9), padx=12, pady=3)
        btn_ejecutar.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        # Botón guardar resultado individual
        btn_guardar = tk.Button(botones_frame, text="💾 Guardar", 
                               command=lambda n=nombre: self.guardar_resultado_individual(n),
                               bg="#27ae60", fg="white", font=("Arial", 9), padx=12, pady=3)
        btn_guardar.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        btn_guardar.config(state=tk.DISABLED)
        
        # Label para resultados texto (cuando hay muchos datos)
        texto_resultado = tk.Label(frame, text="", bg="#16213e", fg="#ecf0f1",
                                   font=("Arial", 8), wraplength=280, justify=tk.LEFT)
        
        self.frames_metodos[nombre] = {
            'frame': frame,
            'canvas': canvas,
            'tiempo_label': tiempo_label,
            'btn_ejecutar': btn_ejecutar,
            'btn_guardar': btn_guardar,
            'texto_resultado': texto_resultado,
            'ultimo_tiempo': None,
            'ultimo_resultado': None,
            'datos_actuales': []
        }
    
    def guardar_resultado_individual(self, nombre):
        """Guarda solo el resultado de un método específico"""
        frame_info = self.frames_metodos[nombre]
        
        if frame_info['ultimo_resultado'] is None:
            messagebox.showwarning("Advertencia", f"Primero ejecute el método '{nombre}'")
            return
        
        from tkinter import filedialog
        
        # Crear nombre de archivo por defecto
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo_default = f"{nombre}_{len(self.datos_actuales)}elem_{fecha_hora}.txt"
        
        archivo = filedialog.asksaveasfilename(
            title=f"Guardar resultado de {nombre}",
            defaultextension=".txt",
            initialfile=nombre_archivo_default,
            filetypes=[
                ("Archivos de texto", "*.txt"),
                ("Archivos CSV", "*.csv"),
                ("Archivos JSON", "*.json"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if archivo:
            try:
                # Determinar formato por extensión
                if archivo.endswith('.json'):
                    self._guardar_como_json(nombre, archivo, frame_info)
                elif archivo.endswith('.csv'):
                    self._guardar_como_csv(nombre, archivo, frame_info)
                else:
                    self._guardar_como_txt(nombre, archivo, frame_info)
                
                self.status_label.config(text=f"✅ Resultado de {nombre} guardado en {archivo.split('/')[-1]}")
                messagebox.showinfo("Éxito", f"Resultado de '{nombre}' guardado correctamente")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")
    
    def _guardar_como_txt(self, nombre, archivo, frame_info):
        """Guarda resultados en formato TXT"""
        with open(archivo, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write(f"RESULTADO DE ORDENAMIENTO - {nombre}\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Método: {nombre}\n")
            f.write(f"Cantidad de elementos: {len(self.datos_actuales)}\n")
            f.write(f"Tiempo de ejecución: {frame_info['ultimo_tiempo']:.6f} segundos\n\n")
            
            f.write("Datos originales:\n")
            if len(self.datos_actuales) <= 100:
                f.write(f"{self.datos_actuales}\n\n")
            else:
                f.write(f"Vista previa (primeros 50): {self.datos_actuales[:50]}\n")
                f.write(f"Total: {len(self.datos_actuales)} elementos\n\n")
            
            f.write("Resultado ordenado:\n")
            if len(frame_info['ultimo_resultado']) <= 100:
                f.write(f"{frame_info['ultimo_resultado']}\n\n")
            else:
                f.write(f"Vista previa (primeros 50): {frame_info['ultimo_resultado'][:50]}\n")
                f.write(f"Vista previa (últimos 50): {frame_info['ultimo_resultado'][-50:]}\n")
                f.write(f"Total: {len(frame_info['ultimo_resultado'])} elementos\n\n")
            
            f.write("-" * 80 + "\n")
    
    def _guardar_como_csv(self, nombre, archivo, frame_info):
        """Guarda resultados en formato CSV"""
        import csv
        with open(archivo, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Metadatos'])
            writer.writerow(['Metodo', nombre])
            writer.writerow(['Fecha', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow(['Cantidad Elementos', len(self.datos_actuales)])
            writer.writerow(['Tiempo (segundos)', f"{frame_info['ultimo_tiempo']:.6f}"])
            writer.writerow([])
            writer.writerow(['Tipo', 'Datos'])
            writer.writerow(['Originales'] + self.datos_actuales)
            writer.writerow(['Ordenados'] + frame_info['ultimo_resultado'])
    
    def _guardar_como_json(self, nombre, archivo, frame_info):
        """Guarda resultados en formato JSON"""
        datos_json = {
            "metadatos": {
                "metodo": nombre,
                "fecha": datetime.now().isoformat(),
                "cantidad_elementos": len(self.datos_actuales),
                "tiempo_segundos": frame_info['ultimo_tiempo']
            },
            "datos_originales": self.datos_actuales,
            "datos_ordenados": frame_info['ultimo_resultado'],
            "tipo_ordenamiento": "interno" if nombre in self.metodos_internos else "externo"
        }
        
        with open(archivo, 'w', encoding='utf-8') as f:
            json.dump(datos_json, f, indent=2, ensure_ascii=False)
    
    def generar_aleatorios(self):
        """Genera datos aleatorios"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Generar Datos")
        dialog.geometry("350x220")
        dialog.configure(bg="#1a1a2e")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        tk.Label(dialog, text="Número de elementos:", bg="#1a1a2e", fg="white", 
                font=("Arial", 10)).pack(pady=10)
        
        var = tk.IntVar(value=15)
        entry = tk.Entry(dialog, textvariable=var, font=("Arial", 10), width=10)
        entry.pack(pady=5)
        
        # Opción para límite máximo
        limite_var = tk.BooleanVar(value=True)
        tk.Checkbutton(dialog, text="Limitar a 30 para visualización gráfica",
                      variable=limite_var, bg="#1a1a2e", fg="white",
                      selectcolor="#1a1a2e", font=("Arial", 9)).pack(pady=5)
        
        def aceptar():
            try:
                n = var.get()
                if n <= 0:
                    messagebox.showerror("Error", "Ingrese un número positivo")
                    return
                
                if n > 1000:
                    if not messagebox.askyesno("Advertencia", f"¿Generar {n} números? Puede ser lento."):
                        return
                
                self.datos_actuales = [random.randint(1, 1000) for _ in range(n)]
                
                # Limpiar resultados anteriores
                self._limpiar_resultados_metodos()
                
                # Mostrar según cantidad
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.nombre_archivo_cargado = ""
                self.status_label.config(text=f"✅ Generados {len(self.datos_actuales)} datos aleatorios")
                
                # Cambiar modo de visualización
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
                if self.modo_visualizacion == "texto":
                    messagebox.showinfo("Modo Texto", f"Se generaron {len(self.datos_actuales)} datos.\nLa visualización será en modo texto (sin gráficos) para mejor rendimiento.\n\nLos resultados se mostrarán como texto y se podrán guardar.")
                
                dialog.destroy()
            except:
                messagebox.showerror("Error", "Ingrese un número válido")
        
        tk.Button(dialog, text="Aceptar", command=aceptar, bg="#27ae60", fg="white", 
                 font=("Arial", 10), padx=20, pady=5).pack(pady=10)
    
    def _limpiar_resultados_metodos(self):
        """Limpia los resultados almacenados de los métodos"""
        for nombre, info in self.frames_metodos.items():
            info['ultimo_resultado'] = None
            info['ultimo_tiempo'] = None
            info['btn_guardar'].config(state=tk.DISABLED)
            info['canvas'].delete("all")
            info['tiempo_label'].config(text="-- s")
            if info['texto_resultado'].winfo_ismapped():
                info['texto_resultado'].pack_forget()
            info['texto_resultado'].config(text="")
            # Asegurar que el canvas está visible
            if not info['canvas'].winfo_ismapped():
                info['canvas'].pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    
    def ingresar_manual(self):
        """Ingreso manual de datos"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Ingresar Datos")
        dialog.geometry("550x350")
        dialog.configure(bg="#1a1a2e")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        tk.Label(dialog, text="Ingrese números separados por espacios o comas:", 
                bg="#1a1a2e", fg="white", font=("Arial", 10)).pack(pady=10)
        
        text_area = tk.Text(dialog, height=10, width=60, font=("Arial", 10), bg="#0f3460", fg="white")
        text_area.pack(pady=10, padx=10)
        
        tk.Label(dialog, text="Ejemplo: 15 8 23 42 16 4", 
                bg="#1a1a2e", fg="#7f8c8d", font=("Arial", 9)).pack()
        
        def aceptar():
            texto = text_area.get("1.0", tk.END)
            texto = texto.replace(',', ' ')
            numeros = texto.split()
            try:
                datos = [int(n) for n in numeros]
                if not datos:
                    messagebox.showerror("Error", "Ingrese al menos un número")
                    return
                
                if len(datos) > 1000:
                    if not messagebox.askyesno("Advertencia", f"Se ingresaron {len(datos)} números. ¿Continuar?"):
                        return
                
                self.datos_actuales = datos
                
                # Limpiar resultados anteriores
                self._limpiar_resultados_metodos()
                
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.nombre_archivo_cargado = ""
                self.status_label.config(text=f"✅ Ingresados {len(datos)} datos manualmente")
                
                # Cambiar modo de visualización
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
                if self.modo_visualizacion == "texto":
                    messagebox.showinfo("Modo Texto", f"Se ingresaron {len(self.datos_actuales)} datos.\nLa visualización será en modo texto (sin gráficos) para mejor rendimiento.\n\nLos resultados se mostrarán como texto y se podrán guardar.")
                
                dialog.destroy()
            except:
                messagebox.showerror("Error", "Ingrese solo números válidos")
        
        btn_frame = tk.Frame(dialog, bg="#1a1a2e")
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Aceptar", command=aceptar, bg="#27ae60", fg="white", padx=20).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancelar", command=dialog.destroy, bg="#e74c3c", fg="white", padx=20).pack(side=tk.LEFT, padx=5)
    
    def cargar_archivo_txt(self):
        """Carga datos desde archivo de texto"""
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo TXT",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    contenido = f.read()
                
                numeros = re.findall(r'-?\d+', contenido)
                datos = [int(n) for n in numeros]
                
                if datos:
                    if len(datos) > 1000:
                        if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos)} números. ¿Continuar?"):
                            return
                    
                    self.datos_actuales = datos
                    
                    # Limpiar resultados anteriores
                    self._limpiar_resultados_metodos()
                    
                    self.nombre_archivo_cargado = archivo
                    
                    if len(self.datos_actuales) <= 30:
                        self.label_datos.config(text=f"Datos ({len(self.datos_actuales)}): {self.datos_actuales}")
                    else:
                        self.label_datos.config(text=f"Datos ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                    
                    self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde {archivo.split('/')[-1]}")
                    
                    self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                    
                    if self.modo_visualizacion == "texto":
                        messagebox.showinfo("Modo Texto", f"Se cargaron {len(self.datos_actuales)} datos.\nLa visualización será en modo texto (sin gráficos) para mejor rendimiento.\n\nLos resultados se mostrarán como texto y se podrán guardar.")
                else:
                    messagebox.showerror("Error", "No se encontraron números en el archivo")
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo: {str(e)}")
    
    def cargar_archivo_json(self):
        """Carga datos desde archivo JSON"""
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo JSON",
            filetypes=[("Archivos JSON", "*.json"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                datos = []
                
                # Intentar extraer números de diferentes estructuras JSON
                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, (int, float)):
                            datos.append(int(item))
                        elif isinstance(item, str) and item.isdigit():
                            datos.append(int(item))
                elif isinstance(data, dict):
                    for key, value in data.items():
                        if isinstance(value, list):
                            for item in value:
                                if isinstance(item, (int, float)):
                                    datos.append(int(item))
                                elif isinstance(item, str) and item.isdigit():
                                    datos.append(int(item))
                        elif isinstance(value, (int, float)):
                            datos.append(int(value))
                else:
                    messagebox.showerror("Error", "Formato JSON no reconocido. Use una lista de números o diccionario con listas.")
                    return
                
                if not datos:
                    messagebox.showerror("Error", "No se encontraron números en el archivo JSON")
                    return
                
                if len(datos) > 1000:
                    if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos)} números. ¿Continuar?"):
                        return
                
                self.datos_actuales = datos
                
                # Limpiar resultados anteriores
                self._limpiar_resultados_metodos()
                
                self.nombre_archivo_cargado = archivo
                
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos JSON ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos JSON ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde {archivo.split('/')[-1]}")
                
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
                if self.modo_visualizacion == "texto":
                    messagebox.showinfo("Modo Texto", f"Se cargaron {len(self.datos_actuales)} datos.\nLa visualización será en modo texto (sin gráficos) para mejor rendimiento.\n\nLos resultados se mostrarán como texto y se podrán guardar.")
                
            except json.JSONDecodeError as e:
                messagebox.showerror("Error", f"Error al decodificar JSON:\n{str(e)}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo: {str(e)}")
    
    def columna_a_indice(self, columna):
        """Convierte letras de columna (A, B, C, AA, AB) a índice numérico (1-indexado)"""
        columna = columna.upper().strip()
        indice = 0
        for char in columna:
            if 'A' <= char <= 'Z':
                indice = indice * 26 + (ord(char) - ord('A') + 1)
            else:
                raise ValueError(f"Columna inválida: {columna}")
        return indice
    
    def cargar_archivo_excel(self):
        """Carga datos desde archivo Excel (.xlsx)"""
        if not EXCEL_DISPONIBLE:
            messagebox.showerror("Error", 
                               "openpyxl no está instalado.\n\n"
                               "Instálalo con el comando:\n"
                               "pip install openpyxl")
            return
        
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            try:
                wb = load_workbook(archivo, data_only=True)
                hojas = wb.sheetnames
                
                if not hojas:
                    messagebox.showerror("Error", "El archivo no contiene hojas")
                    return
                
                self._seleccionar_hoja_excel(archivo, wb, hojas)
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer archivo Excel:\n{str(e)}")
    
    def _seleccionar_hoja_excel(self, archivo, wb, hojas):
        """Diálogo para seleccionar hoja y columna del Excel"""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Seleccionar datos del Excel")
        dialog.geometry("750x650")
        dialog.configure(bg="#1a1a2e")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        # Frame principal con scroll
        main_frame = tk.Frame(dialog, bg="#1a1a2e")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Canvas para scroll del contenido
        canvas = tk.Canvas(main_frame, bg="#1a1a2e", highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#1a1a2e")
        
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Contenido dentro de scroll_frame
        # Selección de hoja
        tk.Label(scroll_frame, text="📍 Seleccione la(s) hoja(s):", bg="#1a1a2e", fg="white",
                font=("Arial", 10, "bold")).pack(pady=5)
        
        # Opción para todas las hojas
        todas_hojas_var = tk.BooleanVar(value=False)
        tk.Checkbutton(scroll_frame, text="Usar TODAS las hojas", variable=todas_hojas_var,
                      bg="#1a1a2e", fg="white", selectcolor="#1a1a2e",
                      command=lambda: toggle_hojas()).pack(pady=5)
        
        # Listbox para seleccionar hojas
        self.hojas_listbox_frame = tk.Frame(scroll_frame, bg="#1a1a2e")
        
        tk.Label(self.hojas_listbox_frame, text="Seleccione hojas específicas (Ctrl+clic para múltiples):", 
                bg="#1a1a2e", fg="white").pack()
        
        self.hojas_listbox = tk.Listbox(self.hojas_listbox_frame, selectmode=tk.MULTIPLE, height=6, width=50)
        self.hojas_listbox.pack(pady=5)
        
        for hoja in hojas:
            self.hojas_listbox.insert(tk.END, hoja)
        
        def toggle_hojas():
            if todas_hojas_var.get():
                self.hojas_listbox_frame.pack_forget()
            else:
                self.hojas_listbox_frame.pack(pady=5)
        
        # Frame para vista previa
        tk.Label(scroll_frame, text="📋 Vista previa de la hoja:", bg="#1a1a2e", fg="white",
                font=("Arial", 10, "bold")).pack(pady=5)
        
        preview_frame = tk.Frame(scroll_frame, bg="#0f3460")
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        preview_text = tk.Text(preview_frame, bg="#0f3460", fg="white", height=10, width=80)
        preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scroll_preview = ttk.Scrollbar(preview_frame, command=preview_text.yview)
        scroll_preview.pack(side=tk.RIGHT, fill=tk.Y)
        preview_text.configure(yscrollcommand=scroll_preview.set)
        
        def actualizar_preview(*args):
            preview_text.delete(1.0, tk.END)
            if not todas_hojas_var.get() and self.hojas_listbox.curselection():
                indices = self.hojas_listbox.curselection()
                for idx in indices:
                    hoja_actual = self.hojas_listbox.get(idx)
                    try:
                        ws = wb[hoja_actual]
                        preview_text.insert(tk.END, f"\n📄 Hoja: {hoja_actual}\n")
                        preview_text.insert(tk.END, "─" * 60 + "\n")
                        
                        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                            if i > 10:
                                preview_text.insert(tk.END, "\n... (más filas no mostradas)")
                                break
                            valores = []
                            for j, cell in enumerate(row[:8]):
                                if cell is not None:
                                    valor_str = str(cell)[:12]
                                    valores.append(f"{chr(65+j)}:{valor_str}")
                            preview_text.insert(tk.END, f"Fila {i:3d}: {' | '.join(valores)}\n")
                    except Exception as e:
                        preview_text.insert(tk.END, f"Error: {str(e)}")
            elif todas_hojas_var.get():
                preview_text.insert(tk.END, "📊 Modo: TODAS LAS HOJAS\n")
                preview_text.insert(tk.END, "─" * 60 + "\n")
                preview_text.insert(tk.END, f"Se cargarán datos de {len(hojas)} hojas\n")
            else:
                preview_text.insert(tk.END, "Seleccione una hoja para ver preview")
        
        self.hojas_listbox.bind('<<ListboxSelect>>', actualizar_preview)
        
        # Selección de columnas
        tk.Label(scroll_frame, text="🔢 Seleccione la(s) columna(s):", bg="#1a1a2e", fg="white",
                font=("Arial", 10, "bold")).pack(pady=5)
        
        col_frame = tk.Frame(scroll_frame, bg="#1a1a2e")
        col_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(col_frame, text="Columnas (ej: A, B, C o A,C,E):", bg="#1a1a2e", fg="white").pack(side=tk.LEFT, padx=5)
        columnas_var = tk.StringVar(value="A")
        columnas_entry = tk.Entry(col_frame, textvariable=columnas_var, width=20)
        columnas_entry.pack(side=tk.LEFT, padx=5)
        
        # Rango de filas
        tk.Label(scroll_frame, text="📊 Rango de filas:", bg="#1a1a2e", fg="white",
                font=("Arial", 10, "bold")).pack(pady=5)
        
        range_frame = tk.Frame(scroll_frame, bg="#1a1a2e")
        range_frame.pack(fill=tk.X, pady=5)
        
        todas_filas_var = tk.BooleanVar(value=True)
        tk.Radiobutton(range_frame, text="Todas las filas", variable=todas_filas_var, value=True,
                      bg="#1a1a2e", fg="white", selectcolor="#1a1a2e").pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(range_frame, text="Rango específico", variable=todas_filas_var, value=False,
                      bg="#1a1a2e", fg="white", selectcolor="#1a1a2e").pack(side=tk.LEFT, padx=10)
        
        self.rango_frame = tk.Frame(scroll_frame, bg="#1a1a2e")
        
        rango_inner = tk.Frame(self.rango_frame, bg="#1a1a2e")
        rango_inner.pack()
        tk.Label(rango_inner, text="Desde fila:", bg="#1a1a2e", fg="white").pack(side=tk.LEFT, padx=5)
        fila_inicio_var = tk.IntVar(value=2)
        tk.Entry(rango_inner, textvariable=fila_inicio_var, width=6).pack(side=tk.LEFT, padx=5)
        tk.Label(rango_inner, text="Hasta fila:", bg="#1a1a2e", fg="white").pack(side=tk.LEFT, padx=5)
        fila_fin_var = tk.IntVar(value=100)
        tk.Entry(rango_inner, textvariable=fila_fin_var, width=6).pack(side=tk.LEFT, padx=5)
        
        def toggle_range_mode():
            if todas_filas_var.get():
                self.rango_frame.pack_forget()
            else:
                self.rango_frame.pack(fill=tk.X, pady=5)
        
        todas_filas_var.trace('w', lambda *args: toggle_range_mode())
        
        # Opción de encabezado
        incluir_encabezado = tk.BooleanVar(value=False)
        tk.Checkbutton(scroll_frame, text="La primera fila es encabezado (ignorar)",
                      variable=incluir_encabezado, bg="#1a1a2e", fg="white",
                      selectcolor="#1a1a2e").pack(pady=5)
        
        def aceptar():
            try:
                # Obtener lista de hojas a procesar
                if todas_hojas_var.get():
                    hojas_a_procesar = hojas
                else:
                    indices = self.hojas_listbox.curselection()
                    if not indices:
                        messagebox.showerror("Error", "Seleccione al menos una hoja")
                        return
                    hojas_a_procesar = [self.hojas_listbox.get(idx) for idx in indices]
                
                # Obtener columnas
                columnas_str = columnas_var.get()
                columnas = [col.strip().upper() for col in columnas_str.split(',')]
                
                # Rango de filas
                if todas_filas_var.get():
                    inicio_fila = 1
                else:
                    inicio_fila = fila_inicio_var.get()
                    if inicio_fila < 1:
                        inicio_fila = 1
                
                datos_totales = []
                
                # Procesar cada hoja
                for hoja in hojas_a_procesar:
                    ws = wb[hoja]
                    
                    if todas_filas_var.get():
                        fin_fila = ws.max_row
                    else:
                        fin_fila = fila_fin_var.get()
                        if fin_fila > ws.max_row:
                            fin_fila = ws.max_row
                    
                    start_row = inicio_fila + 1 if incluir_encabezado.get() else inicio_fila
                    
                    for col in columnas:
                        try:
                            col_idx = self.columna_a_indice(col)
                            for row_idx in range(start_row, fin_fila + 1):
                                cell_value = ws.cell(row=row_idx, column=col_idx).value
                                if cell_value is not None:
                                    try:
                                        num = float(cell_value) if isinstance(cell_value, (int, float)) else int(float(cell_value))
                                        datos_totales.append(int(num))
                                    except (ValueError, TypeError):
                                        pass
                        except Exception as e:
                            messagebox.showwarning("Advertencia", f"Error en hoja {hoja}, columna {col}: {str(e)}")
                
                if not datos_totales:
                    messagebox.showerror("Error", "No se encontraron números en las hojas/columnas seleccionadas")
                    return
                
                if len(datos_totales) > 1000:
                    if not messagebox.askyesno("Advertencia", f"Se encontraron {len(datos_totales)} números. ¿Continuar?"):
                        return
                
                self.datos_actuales = datos_totales
                
                # Limpiar resultados anteriores
                self._limpiar_resultados_metodos()
                
                self.nombre_archivo_cargado = archivo
                
                if len(self.datos_actuales) <= 30:
                    self.label_datos.config(text=f"Datos Excel ({len(self.datos_actuales)}): {self.datos_actuales}")
                else:
                    self.label_datos.config(text=f"Datos Excel ({len(self.datos_actuales)} elementos) - Vista previa: {self.datos_actuales[:20]}...")
                
                self.status_label.config(text=f"✅ Cargados {len(self.datos_actuales)} datos desde {len(hojas_a_procesar)} hoja(s)")
                
                self.modo_visualizacion = "grafico" if len(self.datos_actuales) <= 30 else "texto"
                
                if self.modo_visualizacion == "texto":
                    messagebox.showinfo("Modo Texto", f"Se cargaron {len(self.datos_actuales)} datos.\nLa visualización será en modo texto (sin gráficos) para mejor rendimiento.\n\nLos resultados se mostrarán como texto y se podrán guardar.")
                
                dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar Excel:\n{str(e)}")
        
        # Botones
        btn_frame = tk.Frame(scroll_frame, bg="#1a1a2e")
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="✅ Aceptar", command=aceptar, bg="#27ae60", fg="white",
                 padx=20, pady=5, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="❌ Cancelar", command=dialog.destroy, bg="#e74c3c", fg="white",
                 padx=20, pady=5, font=("Arial", 10)).pack(side=tk.LEFT, padx=10)
    
    def dibujar_barras(self, canvas, datos):
        """Dibuja gráfico de barras en un canvas"""
        canvas.delete("all")
        
        if not datos:
            return
        
        ancho = canvas.winfo_width()
        alto = canvas.winfo_height()
        
        if ancho < 10:
            ancho = 300
        if alto < 10:
            alto = 140
        
        n = len(datos)
        if n == 0:
            return
        
        ancho_barra = max(6, (ancho - 30) / n - 2)
        max_valor = max(datos) if datos else 100
        
        for i, valor in enumerate(datos):
            x0 = 15 + i * (ancho_barra + 2)
            y0 = alto - 25 - (valor / max_valor) * (alto - 50)
            x1 = x0 + ancho_barra
            y1 = alto - 25
            
            # Color basado en el valor (gradiente)
            r = int((valor / max_valor) * 255)
            b = int((1 - valor / max_valor) * 255)
            color = f'#{r:02x}40{b:02x}'
            canvas.create_rectangle(x0, y0, x1, y1, fill=color, outline="white", width=1)
            
            # Mostrar valor si hay espacio suficiente
            if ancho_barra > 12 and n <= 15:
                canvas.create_text(x0 + ancho_barra/2, y0 - 5, text=str(valor),
                                  font=("Arial", 8), fill="white")
    
    def mostrar_resultado_texto(self, frame_info, datos, tiempo):
        """Muestra resultados en modo texto cuando hay muchos datos"""
        frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
        frame_info['ultimo_tiempo'] = tiempo
        frame_info['ultimo_resultado'] = datos
        frame_info['btn_guardar'].config(state=tk.NORMAL)
        
        # Mostrar primeros y últimos elementos
        if len(datos) <= 50:
            texto = f"Resultado: {datos}"
        else:
            texto = f"Resultado ({len(datos)}):\n{datos[:20]}...\n...{datos[-20:]}"
        
        frame_info['texto_resultado'].config(text=texto, font=("Arial", 8))
        
        # Si no está visible, mostrarlo y ocultar canvas
        if not frame_info['texto_resultado'].winfo_ismapped():
            frame_info['canvas'].pack_forget()
            frame_info['texto_resultado'].pack(pady=5, padx=5, fill=tk.BOTH, expand=True)
    
    def ejecutar_individual(self, nombre):
        """Ejecuta un método específico"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "Primero genere o cargue datos")
            return
        
        # Obtener el método
        metodo = None
        if nombre in self.metodos_internos:
            metodo = self.metodos_internos[nombre]
        elif nombre in self.metodos_externos:
            metodo = self.metodos_externos[nombre]
        
        if not metodo:
            return
        
        # Deshabilitar botones mientras ejecuta
        frame_info = self.frames_metodos[nombre]
        frame_info['btn_ejecutar'].config(state=tk.DISABLED, text="Ejecutando...")
        frame_info['btn_guardar'].config(state=tk.DISABLED)
        frame_info['tiempo_label'].config(text="... s")
        
        def ejecutar():
            datos_copy = self.datos_actuales.copy()
            
            start = time.time()
            
            if self.modo_visualizacion == "grafico" and len(datos_copy) <= 30:
                def actualizar_paso(datos_paso):
                    self.dibujar_barras(frame_info['canvas'], datos_paso)
                    frame_info['canvas'].update()
                    time.sleep(0.02)
                
                _, resultado = metodo(datos_copy, actualizar_paso)
            else:
                # Modo texto - sin actualización de pasos
                _, resultado = metodo(datos_copy, None)
            
            tiempo = time.time() - start
            
            # Guardar resultado
            frame_info['ultimo_tiempo'] = tiempo
            frame_info['ultimo_resultado'] = resultado
            
            if self.modo_visualizacion == "grafico" and len(datos_copy) <= 30:
                self.dibujar_barras(frame_info['canvas'], resultado)
                frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
                # Asegurar que canvas está visible
                if frame_info['texto_resultado'].winfo_ismapped():
                    frame_info['texto_resultado'].pack_forget()
                    frame_info['canvas'].pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            else:
                self.mostrar_resultado_texto(frame_info, resultado, tiempo)
            
            frame_info['btn_ejecutar'].config(state=tk.NORMAL, text="▶️ Ejecutar")
            frame_info['btn_guardar'].config(state=tk.NORMAL)
        
        thread = threading.Thread(target=ejecutar)
        thread.daemon = True
        thread.start()
    
    def ejecutar_todos(self):
        """Ejecuta todos los métodos simultáneamente"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "Primero genere o cargue datos")
            return
        
        self.status_label.config(text="🔄 Ejecutando todos los métodos...")
        
        def ejecutar():
            threads = []
            
            for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
                frame_info = self.frames_metodos[nombre]
                frame_info['btn_ejecutar'].config(state=tk.DISABLED, text="Ejecutando...")
                frame_info['btn_guardar'].config(state=tk.DISABLED)
                frame_info['tiempo_label'].config(text="... s")
                
                metodo = None
                if nombre in self.metodos_internos:
                    metodo = self.metodos_internos[nombre]
                else:
                    metodo = self.metodos_externos[nombre]
                
                thread = threading.Thread(target=self._ejecutar_metodo_hilo, args=(nombre, metodo, frame_info))
                thread.daemon = True
                thread.start()
                threads.append(thread)
            
            for thread in threads:
                thread.join()
            
            self.status_label.config(text="✅ Todos los métodos completados")
        
        threading.Thread(target=ejecutar, daemon=True).start()
    
    def _ejecutar_metodo_hilo(self, nombre, metodo, frame_info):
        """Ejecuta un método en un hilo"""
        datos_copy = self.datos_actuales.copy()
        
        start = time.time()
        
        if self.modo_visualizacion == "grafico" and len(datos_copy) <= 30:
            def actualizar_paso(datos_paso):
                self.dibujar_barras(frame_info['canvas'], datos_paso)
                frame_info['canvas'].update()
                time.sleep(0.01)
            
            _, resultado = metodo(datos_copy, actualizar_paso)
        else:
            _, resultado = metodo(datos_copy, None)
        
        tiempo = time.time() - start
        
        # Guardar resultado
        frame_info['ultimo_tiempo'] = tiempo
        frame_info['ultimo_resultado'] = resultado
        
        if self.modo_visualizacion == "grafico" and len(datos_copy) <= 30:
            self.dibujar_barras(frame_info['canvas'], resultado)
            frame_info['tiempo_label'].config(text=f"{tiempo:.3f}s")
            if frame_info['texto_resultado'].winfo_ismapped():
                frame_info['texto_resultado'].pack_forget()
                frame_info['canvas'].pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        else:
            self.mostrar_resultado_texto(frame_info, resultado, tiempo)
        
        frame_info['btn_ejecutar'].config(state=tk.NORMAL, text="▶️ Ejecutar")
        frame_info['btn_guardar'].config(state=tk.NORMAL)
    
    def detener_todos(self):
        """Detiene la ejecución (no puede detener hilos directamente)"""
        self.status_label.config(text="⏹️ Detenido (la ejecución continuará hasta terminar)")
        messagebox.showinfo("Info", "Las ejecuciones en curso continuarán hasta finalizar")
    
    def comparar_rendimiento(self):
        """Compara el rendimiento de todos los métodos"""
        if not self.datos_actuales or len(self.datos_actuales) < 2:
            messagebox.showwarning("Advertencia", "Se necesitan al menos 2 elementos")
            return
        
        resultados = f"""
╔════════════════════════════════════════════════════════════════════════════╗
║                         COMPARACIÓN DE RENDIMIENTO                          ║
╠════════════════════════════════════════════════════════════════════════════╣
║ Elementos: {len(self.datos_actuales):<70} ║
╠════════════════════════════════════════════════════════════════════════════╣
"""
        
        # Medir tiempos para cada método
        for nombre in list(self.metodos_internos.keys()) + list(self.metodos_externos.keys()):
            metodo = None
            if nombre in self.metodos_internos:
                metodo = self.metodos_internos[nombre]
            else:
                metodo = self.metodos_externos[nombre]
            
            datos_copy = self.datos_actuales.copy()
            start = time.time()
            _, _ = metodo(datos_copy)
            tiempo = time.time() - start
            
            resultados += f"║ {nombre:<22} {tiempo:.6f} segundos{' ' * (45 - len(f'{tiempo:.6f}'))}║\n"
        
        resultados += "╚════════════════════════════════════════════════════════════════════════════╝"
        
        # Mostrar resultados en ventana
        ventana_resultados = tk.Toplevel(self.ventana)
        ventana_resultados.title("Resultados de Rendimiento")
        ventana_resultados.geometry("850x650")
        ventana_resultados.configure(bg="#1a1a2e")
        
        text_area = tk.Text(ventana_resultados, bg="#0a0a1a", fg="#00ff00", font=("Courier", 10))
        text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_area.insert(tk.END, resultados)
        text_area.config(state=tk.DISABLED)
        
        btn_frame = tk.Frame(ventana_resultados, bg="#1a1a2e")
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Cerrar", command=ventana_resultados.destroy,
                 bg="#e74c3c", fg="white", padx=20, pady=5).pack()
        
        self.status_label.config(text="📊 Comparación de rendimiento completada")
    
    def guardar_todos_resultados(self):
        """Guarda los resultados de todos los métodos ejecutados"""
        if not self.datos_actuales:
            messagebox.showwarning("Advertencia", "No hay datos para guardar")
            return
        
        # Verificar si hay algún resultado guardado
        resultados_existentes = {nombre: info for nombre, info in self.frames_metodos.items() 
                                if info['ultimo_resultado'] is not None}
        
        if not resultados_existentes:
            messagebox.showwarning("Advertencia", "No hay resultados para guardar. Ejecute al menos un método primero.")
            return
        
        from tkinter import filedialog
        
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo_default = f"todos_resultados_{len(self.datos_actuales)}elem_{fecha_hora}.txt"
        
        archivo = filedialog.asksaveasfilename(
            title="Guardar todos los resultados",
            defaultextension=".txt",
            initialfile=nombre_archivo_default,
            filetypes=[
                ("Archivos de texto", "*.txt"),
                ("Archivos JSON", "*.json"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if archivo:
            try:
                if archivo.endswith('.json'):
                    self._guardar_todos_como_json(archivo, resultados_existentes)
                else:
                    self._guardar_todos_como_txt(archivo, resultados_existentes)
                
                self.status_label.config(text=f"✅ Todos los resultados guardados en {archivo.split('/')[-1]}")
                messagebox.showinfo("Éxito", "Todos los resultados guardados correctamente")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")
    
    def _guardar_todos_como_txt(self, archivo, resultados_existentes):
        """Guarda todos los resultados en formato TXT"""
        with open(archivo, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("REPORTE COMPLETO DE ORDENAMIENTO\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Cantidad de elementos: {len(self.datos_actuales)}\n\n")
            
            f.write("DATOS ORIGINALES:\n")
            if len(self.datos_actuales) <= 100:
                f.write(f"{self.datos_actuales}\n\n")
            else:
                f.write(f"Vista previa (primeros 50): {self.datos_actuales[:50]}\n")
                f.write(f"Total: {len(self.datos_actuales)} elementos\n\n")
            
            f.write("-" * 80 + "\n\n")
            
            # Métodos Internos
            f.write("📊 MÉTODOS INTERNOS\n")
            f.write("=" * 50 + "\n\n")
            for nombre in self.metodos_internos.keys():
                if nombre in resultados_existentes:
                    info = resultados_existentes[nombre]
                    f.write(f"🔸 {nombre}:\n")
                    f.write(f"   Tiempo: {info['ultimo_tiempo']:.6f} segundos\n")
                    if len(info['ultimo_resultado']) <= 100:
                        f.write(f"   Resultado: {info['ultimo_resultado']}\n")
                    else:
                        f.write(f"   Resultado (primeros 50): {info['ultimo_resultado'][:50]}\n")
                    f.write("\n")
                else:
                    f.write(f"🔸 {nombre}: No ejecutado\n\n")
            
            # Métodos Externos
            f.write("\n📊 MÉTODOS EXTERNOS\n")
            f.write("=" * 50 + "\n\n")
            for nombre in self.metodos_externos.keys():
                if nombre in resultados_existentes:
                    info = resultados_existentes[nombre]
                    f.write(f"🔸 {nombre}:\n")
                    f.write(f"   Tiempo: {info['ultimo_tiempo']:.6f} segundos\n")
                    if len(info['ultimo_resultado']) <= 100:
                        f.write(f"   Resultado: {info['ultimo_resultado']}\n")
                    else:
                        f.write(f"   Resultado (primeros 50): {info['ultimo_resultado'][:50]}\n")
                    f.write("\n")
                else:
                    f.write(f"🔸 {nombre}: No ejecutado\n\n")
            
            f.write("=" * 80 + "\n")
    
    def _guardar_todos_como_json(self, archivo, resultados_existentes):
        """Guarda todos los resultados en formato JSON"""
        datos_json = {
            "metadatos": {
                "fecha": datetime.now().isoformat(),
                "cantidad_elementos": len(self.datos_actuales)
            },
            "datos_originales": self.datos_actuales,
            "resultados": {}
        }
        
        for nombre, info in resultados_existentes.items():
            datos_json["resultados"][nombre] = {
                "tiempo_segundos": info['ultimo_tiempo'],
                "datos_ordenados": info['ultimo_resultado'],
                "tipo": "interno" if nombre in self.metodos_internos else "externo"
            }
        
        with open(archivo, 'w', encoding='utf-8') as f:
            json.dump(datos_json, f, indent=2, ensure_ascii=False)
    
    def limpiar_todo(self):
        """Limpia todos los datos y resultados"""
        if messagebox.askyesno("Confirmar", "¿Está seguro de limpiar todos los datos y resultados?"):
            self.datos_actuales = []
            self.label_datos.config(text="Sin datos")
            self.nombre_archivo_cargado = ""
            self._limpiar_resultados_metodos()
            self.status_label.config(text="✅ Todo limpiado")
    
    def ejecutar(self):
        """Inicia la aplicación"""
        self.ventana.mainloop()


if __name__ == "__main__":
    app = DemoCompletaOrdenamiento()
    app.ejecutar()